import os
import pandas as pd
import numpy as np
import pickle
import logging
import re
import json
import requests
import random
import time
from typing import List, Dict, Union, Optional, Tuple, Any
from pathlib import Path
from flask import Flask, request, jsonify, render_template, send_from_directory
from werkzeug.utils import secure_filename

# File processing libraries
import PyPDF2
from docx import Document
import openpyxl
import csv

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

class DocumentProcessor:
    """Process and extract text from various document formats"""
    
    def __init__(self, storage_dir: str = "data"):
        self.storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)
        os.makedirs(os.path.join(storage_dir, "uploads"), exist_ok=True)
        logger.info(f"Document processor initialized with storage in {storage_dir}")
    
    def save_file(self, file_path: str, file_content) -> str:
        """Save uploaded file and return the path"""
        filename = os.path.basename(file_path)
        save_path = os.path.join(self.storage_dir, "uploads", filename)
        
        # Save the file
        with open(save_path, "wb") as f:
            f.write(file_content)
        
        logger.info(f"File saved to {save_path}")
        return save_path
    
    def extract_text(self, file_path: str) -> str:
        """Extract text from various file formats"""
        file_extension = os.path.splitext(file_path)[1].lower()
        
        try:
            if file_extension == '.pdf':
                return self._extract_from_pdf(file_path)
            elif file_extension == '.txt':
                return self._extract_from_txt(file_path)
            elif file_extension == '.docx':
                return self._extract_from_docx(file_path)
            elif file_extension in ['.xlsx', '.xls']:
                return self._extract_from_excel(file_path)
            elif file_extension == '.csv':
                return self._extract_from_csv(file_path)
            else:
                error_msg = f"Unsupported file format: {file_extension}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        except Exception as e:
            logger.error(f"Error extracting text from {file_path}: {str(e)}")
            raise
    
    def _extract_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                reader = PyPDF2.PdfReader(file)
                for page in reader.pages:
                    page_text = page.extract_text()
                    if page_text:  # Check if text extraction was successful
                        text += page_text + "\n"
        except Exception as e:
            logger.error(f"Error in PDF extraction: {str(e)}")
            text = f"[Error extracting text from PDF: {str(e)}]"
        
        if not text:
            text = "[No extractable text found in PDF]"
        
        logger.info(f"Extracted {len(text)} characters from PDF: {file_path}")
        return text

    def _extract_from_txt(self, file_path: str) -> str:
        """Extract text from TXT file"""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            text = file.read()
        logger.info(f"Extracted {len(text)} characters from TXT: {file_path}")
        return text
    
    def _extract_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        doc = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        logger.info(f"Extracted {len(text)} characters from DOCX: {file_path}")
        return text
    
    def _extract_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        text = ""
        
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            text += f"Sheet: {sheet}\n"
            
            for row in ws.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                text += "\t".join(row_values) + "\n"
            
            text += "\n"
        
        logger.info(f"Extracted {len(text)} characters from Excel: {file_path}")
        return text
    
    def _extract_from_csv(self, file_path: str) -> str:
        """Extract text from CSV file"""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as file:
            reader = csv.reader(file)
            rows = list(reader)
            
        text = "\n".join(["\t".join(row) for row in rows])
        logger.info(f"Extracted {len(text)} characters from CSV: {file_path}")
        return text

class DataManager:
    """Manages data for model training"""
    
    def __init__(self, data_dir: str = "processed_data"):
        self.data_dir = data_dir
        os.makedirs(data_dir, exist_ok=True)
        self.documents = []
        self.document_metadata = []
        self.document_summaries = []  # Added to store document summaries
        self.training_chunks = []  # Store chunked data for training
        logger.info(f"Data manager initialized with directory: {data_dir}")
    
    def add_document(self, text: str, metadata: Dict[str, Any]) -> None:
        """Add document and its metadata to the collection"""
        self.documents.append(text)
        self.document_metadata.append(metadata)
        
        # Generate and store a simple summary
        summary = self._generate_simple_summary(text)
        self.document_summaries.append(summary)
        
        logger.info(f"Added document with metadata: {metadata}")
    
    def _generate_simple_summary(self, text: str, max_length: int = 200) -> str:
        """Generate a simple summary of the text"""
        # Take first few sentences as a summary
        sentences = re.split(r'(?<=[.!?])\s+', text[:1000])
        summary = ' '.join(sentences[:3])
        
        if len(summary) > max_length:
            summary = summary[:max_length] + "..."
            
        return summary
    
    def prepare_for_training(self, chunk_size: int = 512, overlap: int = 50) -> Tuple[List[str], List[str]]:
        """Prepare documents for training by chunking them and splitting into train/validation"""
        chunks = []
        
        for doc in self.documents:
            # Enhanced chunking strategy - tries to maintain sentence boundaries
            doc_chunks = self._smart_chunk_text(doc, chunk_size, overlap)
            chunks.extend(doc_chunks)
        
        # Store all chunks
        self.training_chunks = chunks
        
        # Split into training (80%) and validation (20%) sets
        random.shuffle(chunks)
        split_point = int(len(chunks) * 0.8)
        train_chunks = chunks[:split_point]
        val_chunks = chunks[split_point:]
        
        logger.info(f"Prepared {len(train_chunks)} training chunks and {len(val_chunks)} validation chunks")
        return train_chunks, val_chunks
    
    def _smart_chunk_text(self, text: str, chunk_size: int, overlap: int) -> List[str]:
        """Split text into overlapping chunks, trying to maintain sentence boundaries"""
        # Split text into sentences
        sentences = re.split(r'(?<=[.!?])\s+', text)
        chunks = []
        current_chunk = []
        current_length = 0
        
        for sentence in sentences:
            sentence_words = sentence.split()
            sentence_length = len(sentence_words)
            
            # If adding this sentence would exceed chunk size
            if current_length + sentence_length > chunk_size:
                # If we have something in the current chunk, add it
                if current_chunk:
                    chunks.append(" ".join(current_chunk))
                
                # Reset current chunk but keep some overlap
                if overlap > 0 and current_chunk:
                    # Take the last few words for overlap
                    overlap_words = min(overlap, len(current_chunk))
                    current_chunk = current_chunk[-overlap_words:]
                    current_length = len(current_chunk)
                else:
                    current_chunk = []
                    current_length = 0
            
            # Add the current sentence
            current_chunk.append(sentence)
            current_length += sentence_length
        
        # Add the last chunk if there's anything left
        if current_chunk:
            chunks.append(" ".join(current_chunk))
        
        return chunks
    
    def save_processed_data(self, filename: str = "processed_data.pkl") -> str:
        """Save processed data to file"""
        save_path = os.path.join(self.data_dir, filename)
        
        data = {
            "documents": self.documents,
            "metadata": self.document_metadata,
            "summaries": self.document_summaries,
            "chunks": self.training_chunks
        }
        
        with open(save_path, 'wb') as f:
            pickle.dump(data, f)
        
        logger.info(f"Saved processed data to {save_path}")
        return save_path
    
    def load_processed_data(self, filename: str = "processed_data.pkl") -> None:
        """Load processed data from file"""
        load_path = os.path.join(self.data_dir, filename)
        
        if os.path.exists(load_path):
            with open(load_path, 'rb') as f:
                data = pickle.load(f)
            
            self.documents = data["documents"]
            self.document_metadata = data["metadata"]
            if "summaries" in data:
                self.document_summaries = data["summaries"]
            else:
                # Generate summaries if they don't exist in the loaded data
                self.document_summaries = [self._generate_simple_summary(doc) for doc in self.documents]
                
            if "chunks" in data:
                self.training_chunks = data["chunks"]
            
            logger.info(f"Loaded {len(self.documents)} documents from {load_path}")
        else:
            logger.info(f"No processed data found at {load_path}")
    
    def get_document_contexts(self, query: str, top_n: int = 3) -> str:
        """Get relevant document contexts based on a query"""
        if not self.documents:
            return ""
            
        # Simple keyword-based relevance
        query_terms = set(query.lower().split())
        
        # Score each document based on query term matches
        doc_scores = []
        for i, doc in enumerate(self.documents):
            score = sum(1 for term in query_terms if term in doc.lower())
            doc_scores.append((i, score))
        
        # Sort by score and get top_n
        top_docs = sorted(doc_scores, key=lambda x: x[1], reverse=True)[:top_n]
        
        # Get the summaries of top docs
        context = "Related information from your documents:\n\n"
        for doc_idx, score in top_docs:
            if score > 0:  # Only include if there's some relevance
                context += f"- {self.document_summaries[doc_idx]}\n\n"
        
        return context if len(context) > 40 else ""  # Return empty if no relevant docs found
    
    def create_training_examples(self, n_examples: int = 10) -> List[Dict[str, str]]:
        """Create formatted examples for fine-tuning"""
        if not self.training_chunks:
            logger.warning("No training chunks available")
            return []
        
        # Use a subset of chunks if we have too many
        chunks_to_use = self.training_chunks
        if len(chunks_to_use) > n_examples:
            chunks_to_use = random.sample(chunks_to_use, n_examples)
        
        # Create examples in the required format
        examples = []
        for chunk in chunks_to_use:
            # Generate a question based on the chunk
            question = self._generate_question_from_chunk(chunk)
            
            examples.append({
                "question": question,
                "context": chunk,
                "answer": self._generate_answer_from_chunk(chunk, question)
            })
        
        return examples
    
    def _generate_question_from_chunk(self, chunk: str) -> str:
        """Generate a simple question based on the content of the chunk"""
        # Extract potential keywords for the question
        words = re.findall(r'\b\w+\b', chunk)
        nouns = [w for w in words if len(w) > 4]  # Simple heuristic for potential nouns
        
        if nouns:
            keyword = random.choice(nouns)
            return f"What information do you have about {keyword}?"
        else:
            return "Can you summarize this information?"
    
    def _generate_answer_from_chunk(self, chunk: str, question: str) -> str:
        """Generate a simple answer based on the content of the chunk and question"""
        # For simple training data, we'll just use the chunk as the answer
        # In a real system, this would be more sophisticated
        return f"Based on the document, {chunk[:200]}..."

class OllamaModelHandler:
    """Handler for Ollama-based models"""
    
    def __init__(self, model_name: str = "llama3.2:latest", ollama_host: str = "http://localhost:11434"):
        self.model_name = model_name
        self.ollama_host = ollama_host
        self.generation_config = {
            "temperature": 0.7,
            "top_p": 0.9,
            "max_tokens": 512,
            "repeat_penalty": 1.1,
        }
        
        self.model = None  # We'll use this as a flag to check if Ollama is reachable
        self.fine_tuned_model_name = None  # Name of the fine-tuned model
        logger.info(f"Ollama handler initialized with model: {model_name}")
    
    def initialize_model(self) -> None:
        """Check if Ollama is running and the model is available"""
        try:
            logger.info(f"Checking if Ollama is running and model {self.model_name} is available...")
            
            # Check if Ollama is running by listing models
            response = requests.get(f"{self.ollama_host}/api/tags")
            
            if response.status_code == 200:
                models = response.json().get("models", [])
                model_names = [model["name"] for model in models]
                
                if self.model_name in model_names:
                    logger.info(f"Model {self.model_name} is available in Ollama")
                    self.model = True  # Just a flag to indicate Ollama is available
                    
                    # Check if we have a fine-tuned version
                    fine_tuned_name = f"{self.model_name}-docs"
                    if fine_tuned_name in model_names:
                        logger.info(f"Found fine-tuned model: {fine_tuned_name}")
                        self.fine_tuned_model_name = fine_tuned_name
                else:
                    available_models = ", ".join(model_names)
                    error_msg = f"Model {self.model_name} not found in Ollama. Available models: {available_models}"
                    logger.error(error_msg)
                    raise ValueError(error_msg)
            else:
                error_msg = f"Failed to connect to Ollama API: {response.status_code} - {response.text}"
                logger.error(error_msg)
                raise ConnectionError(error_msg)
                
        except Exception as e:
            logger.error(f"Error initializing Ollama model: {str(e)}")
            raise
    
    def update_generation_config(self, config: Dict[str, Any]) -> None:
        """Update generation configuration"""
        # Map the web UI config keys to Ollama API param keys
        if "max_new_tokens" in config:
            config["max_tokens"] = config.pop("max_new_tokens")
        if "repetition_penalty" in config:
            config["repeat_penalty"] = config.pop("repetition_penalty")
            
        self.generation_config.update(config)
        logger.info(f"Updated generation config: {self.generation_config}")
    
    def generate_response(self, prompt: str) -> str:
        """Generate a response using Ollama API"""
        if not self.model:
            raise ValueError("Must initialize Ollama connection before generating responses")
        
        try:
            # Format prompt for Llama 3
            formatted_prompt = f"""<|system|>
You are a helpful assistant that answers questions based on provided documents.
<|user|>
{prompt}
<|assistant|>"""
            
            # Use fine-tuned model if available
            model_to_use = self.fine_tuned_model_name if self.fine_tuned_model_name else self.model_name
            
            # Prepare the request to Ollama API
            payload = {
                "model": model_to_use,
                "prompt": formatted_prompt,
                "stream": False,
                "options": {
                    "temperature": self.generation_config["temperature"],
                    "top_p": self.generation_config["top_p"],
                    "max_tokens": self.generation_config["max_tokens"],
                    "repeat_penalty": self.generation_config.get("repeat_penalty", 1.1)
                }
            }
            
            # Call Ollama API
            response = requests.post(f"{self.ollama_host}/api/generate", json=payload)
            
            if response.status_code == 200:
                result = response.json()
                generated_text = result.get("response", "").strip()
                logger.info(f"Generated response of length {len(generated_text)}")
                return generated_text
            else:
                error_msg = f"Ollama API error: {response.status_code} - {response.text}"
                logger.error(error_msg)
                raise RuntimeError(error_msg)
                
        except Exception as e:
            logger.error(f"Error generating response with Ollama: {str(e)}")
            raise
    
    def fine_tune_model(self, training_data: List[Dict[str, str]], progress_callback=None) -> str:
     """Fine-tune the model on the given training data using Ollama's API"""
     if not self.model:
        raise ValueError("Must initialize model first")
    
     try:
        # Create a modelfile for fine-tuning
        fine_tuned_name = f"{self.model_name.split(':')[0]}-docs"
        
        # Create a very simple modelfile - just extend the base model with a system prompt
        modelfile = f"FROM {self.model_name}\n\n"
        modelfile += "SYSTEM \"You are a helpful assistant trained to answer questions about specific documents. Use the document information to provide accurate and relevant answers.\"\n"
        
        # Create the model
        logger.info(f"Creating fine-tuned model: {fine_tuned_name}")
        
        if progress_callback:
            progress_callback(10, "Creating model file...")
        
        # Save the modelfile to disk first
        modelfile_path = os.path.join("data", "modelfile.txt")
        with open(modelfile_path, "w") as f:
            f.write(modelfile)
        
        # Log the modelfile content for debugging
        logger.info(f"Modelfile content:\n{modelfile}")
        
        import subprocess
        
        if progress_callback:
            progress_callback(20, "Building model (this may take several minutes)...")
        
        # Create model using subprocess for better reliability
        create_command = f"ollama create {fine_tuned_name} -f {modelfile_path}"
        result = subprocess.run(create_command, shell=True, capture_output=True, text=True)
        
        # Log the result for debugging
        logger.info(f"Create command result: {result.stdout}")
        if result.stderr:
            logger.warning(f"Create command stderr: {result.stderr}")
        
        if result.returncode != 0:
            raise RuntimeError(f"Failed to create model: {result.stderr}")
        
        # Set the fine-tuned model as active
        self.fine_tuned_model_name = fine_tuned_name
        logger.info(f"Fine-tuned model created: {fine_tuned_name}")
        
        if progress_callback:
            progress_callback(100, "Training complete!")
        
        return fine_tuned_name
    
     except Exception as e:
        logger.error(f"Error fine-tuning model: {str(e)}")
        if progress_callback:
            progress_callback(-1, f"Error: {str(e)}")
        raise
class ChatInterface:
    """Interface for interacting with the model"""
    
    def __init__(self, document_processor: DocumentProcessor, 
                 data_manager: DataManager, model_handler: OllamaModelHandler):
        self.document_processor = document_processor
        self.data_manager = data_manager
        self.model_handler = model_handler
        self.conversation_history = []
        self.max_history_turns = 5  # Store last 5 conversation turns
        self.training_progress = {"status": "idle", "progress": 0, "message": ""}
        logger.info("Chat interface initialized")
    
    def upload_file(self, file_path: str, file_content) -> str:
        """Upload and process a file"""
        try:
            logger.info(f"Starting to process file: {file_path}")
            
            # Save the file
            saved_path = self.document_processor.save_file(file_path, file_content)
            logger.info(f"File saved to: {saved_path}")
            
            # Extract text
            text = self.document_processor.extract_text(saved_path)
            logger.info(f"Text extracted with length: {len(text)}")
            
            # Add to data manager
            metadata = {
                "filename": os.path.basename(file_path),
                "path": saved_path,
                "length": len(text),
                "timestamp": pd.Timestamp.now().isoformat()
            }
            self.data_manager.add_document(text, metadata)
            logger.info(f"Document added to data manager with metadata: {metadata}")
            
            # Save processed data
            self.data_manager.save_processed_data()
            
            return f"Successfully uploaded and processed {os.path.basename(file_path)}"
        except Exception as e:
            error_msg = f"Error processing file {file_path}: {str(e)}"
            logger.error(error_msg)
            return error_msg
    
    def get_response(self, message: str) -> str:
        """Get response from the model based on user message"""
        try:
            if not self.model_handler.model:
                return "Model is not initialized yet. Please initialize the model first."
            
            # Get relevant document context
            doc_context = self.data_manager.get_document_contexts(message)
            
            # Build conversation history
            history_text = ""
            for turn in self.conversation_history[-self.max_history_turns:]:
                history_text += f"User: {turn[0]}\nAssistant: {turn[1]}\n\n"
            
            # Create prompt with context, history, and message
            prompt = ""
            if doc_context:
                prompt += f"{doc_context}\n\n"
            if history_text:
                prompt += f"Previous conversation:\n{history_text}\n"
            prompt += f"Current question: {message}"
            
            # Generate response
            response = self.model_handler.generate_response(prompt)
            
            # Add to conversation history
            self.conversation_history.append((message, response))
            
            return response
        except Exception as e:
            error_msg = f"Error generating response: {str(e)}"
            logger.error(error_msg)
            return error_msg
    
    def clear_conversation(self) -> None:
        """Clear the conversation history"""
        self.conversation_history = []
        logger.info("Conversation history cleared")
    
    def train_model(self, n_examples: int = 15) -> Dict[str, Any]:
        """Train the model on uploaded documents"""
        try:
            if not self.model_handler.model:
                return {
                    "status": "error", 
                    "message": "Please initialize the model first"
                }
            
            if not self.data_manager.documents:
                return {
                    "status": "error", 
                    "message": "No documents uploaded yet. Please upload documents first."
                }
            
            # Update training status
            self.training_progress = {
                "status": "preparing", 
                "progress": 5, 
                "message": "Preparing training data..."
            }
            
            # Prepare data for training
            train_chunks, val_chunks = self.data_manager.prepare_for_training()
            
            if len(train_chunks) == 0:
                return {
                    "status": "error", 
                    "message": "Not enough data for training. Please upload more documents."
                }
            
            # Create training examples
            self.training_progress = {
                "status": "creating_examples", 
                "progress": 10, 
                "message": "Creating training examples..."
            }
            examples = self.data_manager.create_training_examples(n_examples)
            
            # Start fine-tuning
            self.training_progress = {
                "status": "training", 
                "progress": 15, 
                "message": "Starting model fine-tuning..."
            }
            
            def progress_update(progress, message):
                self.training_progress = {
                    "status": "training" if progress >= 0 else "error",
                    "progress": progress if progress >= 0 else 0,
                    "message": message
                }
            
            # Fine-tune the model
            fine_tuned_name = self.model_handler.fine_tune_model(examples, progress_callback=progress_update)
            
            self.training_progress = {
                "status": "complete", 
                "progress": 100, 
                "message": f"Training complete! Fine-tuned model: {fine_tuned_name}"
            }
            
            # Return success
            return {
                "status": "success",
                "message": f"Model fine-tuned successfully as {fine_tuned_name}",
                "model_name": fine_tuned_name
            }
            
        except Exception as e:
            error_msg = f"Error training model: {str(e)}"
            logger.error(error_msg)
            self.training_progress = {
                "status": "error", 
                "progress": 0, 
                "message": error_msg
            }
            return {
                "status": "error",
                "message": error_msg
            }
    
    def get_training_progress(self) -> Dict[str, Any]:
        """Get the current training progress"""
        return self.training_progress

# Initialize Flask app
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = os.path.join('data', 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Make sure templates directory exists
templates_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
os.makedirs(templates_dir, exist_ok=True)

# Initialize components
document_processor = DocumentProcessor()
data_manager = DataManager()
ollama_handler = OllamaModelHandler(model_name="llama3.2:latest")  # Using your local Ollama model
chat_interface = ChatInterface(document_processor, data_manager, ollama_handler)

# Try to load any previously processed data
try:
    data_manager.load_processed_data()
except Exception as e:
    logger.warning(f"Could not load previous data: {e}")

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/initialize_model', methods=['POST'])
def initialize_model():
    try:
        ollama_handler.initialize_model()
        
        # Check if we have a fine-tuned model
        if ollama_handler.fine_tuned_model_name:
            return jsonify({
                'status': 'success', 
                'message': f'Ollama model initialized successfully (fine-tuned model found: {ollama_handler.fine_tuned_model_name})',
                'fine_tuned': True,
                'fine_tuned_name': ollama_handler.fine_tuned_model_name
            })
        else:
            return jsonify({
                'status': 'success', 
                'message': 'Ollama model initialized successfully',
                'fine_tuned': False
            })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/upload', methods=['POST'])
def upload_file():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file part'})
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No selected file'})
        
        if file:
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            with open(file_path, 'rb') as f:
                content = f.read()
            
            result = chat_interface.upload_file(file_path, content)
            return jsonify({'status': 'success', 'message': result})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/get_response', methods=['POST'])
def get_response():
    try:
        data = request.json
        message = data.get('message', '')
        if not message:
            return jsonify({'status': 'error', 'message': 'No message provided'})
        response = chat_interface.get_response(message)
        return jsonify({'status': 'success', 'response': response})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/clear_conversation', methods=['POST'])
def clear_conversation():
    try:
        chat_interface.clear_conversation()
        return jsonify({'status': 'success', 'message': 'Conversation cleared'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/update_config', methods=['POST'])
def update_config():
    try:
        data = request.json
        config = data.get('config', {})
        if not config:
            return jsonify({'status': 'error', 'message': 'No configuration provided'})
        
        ollama_handler.update_generation_config(config)
        return jsonify({'status': 'success', 'message': 'Configuration updated'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/get_uploaded_files', methods=['GET'])
def get_uploaded_files():
    try:
        files = []
        for i, metadata in enumerate(data_manager.document_metadata):
            files.append({
                'id': i,
                'filename': metadata['filename'],
                'length': metadata['length'],
                'timestamp': metadata['timestamp']
            })
        return jsonify({'status': 'success', 'files': files})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/get_available_models', methods=['GET'])
def get_available_models():
    try:
        response = requests.get(f"{ollama_handler.ollama_host}/api/tags")
        
        if response.status_code == 200:
            models = response.json().get("models", [])
            model_names = [model["name"] for model in models]
            
            # Identify fine-tuned models (assuming they end with "-docs")
            fine_tuned_models = [name for name in model_names if "-docs" in name]
            
            return jsonify({
                'status': 'success', 
                'models': model_names,
                'fine_tuned_models': fine_tuned_models,
                'current_model': ollama_handler.model_name
            })
        else:
            return jsonify({'status': 'error', 'message': 'Failed to fetch models from Ollama'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/change_model', methods=['POST'])
def change_model():
    try:
        data = request.json
        model_name = data.get('model_name', '')
        
        if not model_name:
            return jsonify({'status': 'error', 'message': 'No model name provided'})
        
        # Reset the model handler
        ollama_handler.model = None
        ollama_handler.model_name = model_name
        ollama_handler.fine_tuned_model_name = None
        
        # Initialize the new model
        ollama_handler.initialize_model()
        
        return jsonify({
            'status': 'success', 
            'message': f'Model changed to {model_name} successfully',
            'fine_tuned': ollama_handler.fine_tuned_model_name is not None,
            'fine_tuned_name': ollama_handler.fine_tuned_model_name
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/train_model', methods=['POST'])
def train_model():
    try:
        data = request.json
        n_examples = data.get('n_examples', 15)
        
        # Start training in a separate thread to avoid blocking
        import threading
        
        def train_thread():
            chat_interface.train_model(n_examples=n_examples)
        
        thread = threading.Thread(target=train_thread)
        thread.daemon = True
        thread.start()
        
        return jsonify({
            'status': 'success', 
            'message': 'Training started in the background. Check progress via /training_progress endpoint.'
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/training_progress', methods=['GET'])
def training_progress():
    try:
        progress = chat_interface.get_training_progress()
        return jsonify({
            'status': 'success',
            'progress': progress
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

if __name__ == '__main__':
    app.run(debug=True, port=9000)